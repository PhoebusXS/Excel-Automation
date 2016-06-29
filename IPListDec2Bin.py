from openpyxl import load_workbook

def ipDec2Bin(ip):
	ansList = []
	digList = ip.split('.')
	for dig in digList:
		binDig = bin(int(dig))
		ansList.append(binDig[2:len(binDig)])
	return '.'.join(ansList)

wb = load_workbook('ip.xlsx')
ws = wb['Sheet 1']
for col in ws.columns[0:1]:
	for cell in col:
		if cell.value:
			try:
				ipBin = ipDec2Bin(cell.value)
				ws['B%d' % int(cell.coordinate[1:len(cell.coordinate)])] = ipBin
			except:
				pass

wb.save('resultIPList.xlsx')
